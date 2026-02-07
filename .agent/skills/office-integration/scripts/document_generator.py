#!/usr/bin/env python3
"""
Document Generator Script

Generate documents from templates and data.
Usage: python document_generator.py <template> <data> --output <output>

Examples:
    python document_generator.py template.docx data.json --output report.docx
    python document_generator.py template.xlsx data.csv --output report.xlsx
"""

from __future__ import annotations

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parents[3] / "scripts"))
from _console import console, success, error, warning, step

import argparse
import json
from dataclasses import dataclass
from enum import Enum
from typing import Any, Dict, List


class TemplateType(str, Enum):
    """Supported template types."""
    WORD = "docx"
    EXCEL = "xlsx"


@dataclass
class GenerationResult:
    """Result of document generation."""
    success: bool
    output_path: Path | None
    message: str
    records_processed: int = 0


def load_data(path: Path) -> List[Dict[str, Any]]:
    """Load data from JSON or CSV file."""
    suffix = path.suffix.lower()
    
    if suffix == ".json":
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, dict):
            return [data]
        return data
    
    elif suffix == ".csv":
        import csv
        with open(path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            return list(reader)
    
    else:
        raise ValueError(f"Unsupported data format: {suffix}")


def generate_word(
    template_path: Path,
    data: Dict[str, Any],
    output_path: Path
) -> GenerationResult:
    """Generate Word document from template."""
    try:
        from docxtpl import DocxTemplate
    except ImportError:
        return GenerationResult(
            success=False,
            output_path=None,
            message="docxtpl not installed. Run: pip install docxtpl"
        )
    
    try:
        doc = DocxTemplate(template_path)
        doc.render(data)
        doc.save(output_path)
        
        return GenerationResult(
            success=True,
            output_path=output_path,
            message=f"Generated: {output_path}",
            records_processed=1
        )
    except Exception as e:
        return GenerationResult(
            success=False,
            output_path=None,
            message=f"Generation failed: {e}"
        )


def generate_excel(
    template_path: Path,
    data: List[Dict[str, Any]],
    output_path: Path
) -> GenerationResult:
    """Generate Excel from template with data."""
    try:
        import openpyxl
    except ImportError:
        return GenerationResult(
            success=False,
            output_path=None,
            message="openpyxl not installed. Run: pip install openpyxl"
        )
    
    try:
        if template_path.exists():
            wb = openpyxl.load_workbook(template_path)
        else:
            wb = openpyxl.Workbook()
        
        ws = wb.active
        
        if data:
            # Write headers
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Write data
            for row_idx, record in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=record.get(header))
        
        wb.save(output_path)
        
        return GenerationResult(
            success=True,
            output_path=output_path,
            message=f"Generated: {output_path}",
            records_processed=len(data)
        )
    except Exception as e:
        return GenerationResult(
            success=False,
            output_path=None,
            message=f"Generation failed: {e}"
        )


def print_result(result: GenerationResult) -> None:
    """Print generation result."""
    if result.success:
        success(f"{result.message} ({result.records_processed} records)")
    else:
        error(result.message)


def main() -> int:
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description="Document Generator - Create documents from templates",
    )
    parser.add_argument("template", type=Path, help="Template file path")
    parser.add_argument("data", type=Path, help="Data file (JSON or CSV)")
    parser.add_argument(
        "--output", "-o",
        type=Path,
        required=True,
        help="Output file path"
    )
    parser.add_argument(
        "--batch",
        action="store_true",
        help="Generate one document per data record"
    )
    
    args = parser.parse_args()
    
    # Validate inputs
    if not args.template.exists():
        print(f"ERROR: Template not found: {args.template}")
        return 1
    
    if not args.data.exists():
        print(f"ERROR: Data file not found: {args.data}")
        return 1
    
    # Load data
    try:
        data = load_data(args.data)
    except Exception as e:
        print(f"ERROR: Failed to load data: {e}")
        return 1
    
    # Determine template type
    template_type = args.template.suffix.lower().lstrip(".")
    
    # Generate
    if template_type == "docx":
        if args.batch and len(data) > 1:
            # Batch mode: one document per record
            output_dir = args.output.parent
            output_dir.mkdir(parents=True, exist_ok=True)
            
            for i, record in enumerate(data):
                output_path = output_dir / f"{args.output.stem}_{i+1}{args.output.suffix}"
                result = generate_word(args.template, record, output_path)
                print_result(result)
        else:
            # Single document
            result = generate_word(args.template, data[0] if data else {}, args.output)
            print_result(result)
    
    elif template_type == "xlsx":
        result = generate_excel(args.template, data, args.output)
        print_result(result)
    
    else:
        print(f"ERROR: Unsupported template type: {template_type}")
        return 1
    
    return 0 if result.success else 1


if __name__ == "__main__":
    sys.exit(main())
